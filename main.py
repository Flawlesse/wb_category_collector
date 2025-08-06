import aiohttp
import asyncio
import openpyxl
from collections import deque
import logging
import time


logging.basicConfig()
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)


async def parse_filters(cat):
    url = (
        "https://catalog.wb.ru/catalog/{shard}/v8/filters?"
        + "ab_testing=false&appType=1&{query}&curr=byn"
        + "&dest=-1257786&hide_dtype=10;13;14&lang=ru"
    )

    name = cat["name"]
    shard = cat["shard"]
    cat_id = cat["id"]
    query = cat.get("query", f"cat={cat_id}")  # this is important

    url = url.format(shard=shard, query=query)

    logger.info(
        f"Parsing list category name='{name}' id={cat_id} {shard=} "
        + f"{query=}...\nQuery: {url}"
    )

    if shard == "blackhole":
        logger.warning(
            f"List category name='{name}' id={cat_id} {shard=} {query=} "
            + "is a blackhole shard. Error on WB side, aborting."
        )
        return []

    # just a category, NOT a subject extension
    async with aiohttp.ClientSession() as session:
        async with session.get(url) as response:
            if response.status >= 400:
                logger.error(
                    f"Error processing {url=}: status={response.status}"
                )
            data = await response.json()
            filters = data["data"]["filters"]
            cat_filter = [f for f in filters if f["name"] == "Категория"]
            if not cat_filter:
                logger.warning(
                    f"List category name='{name}' id={cat_id} {shard=} "
                    + f"{query=} does not have 'Категория' filter."
                )
                # check if category is subject / set of subjects itself
                if "subject" in query:
                    subjects = query[8:].split(";")
                    if len(subjects) > 1:
                        logger.warning(
                            f"List category name='{name}' id={cat_id} {shard=} "
                            + f"{query=} has multiple subjects. Unprocessable."
                        )
                        return []

                    return [[subjects[0], name, 99]]

                return []

            cat_filter = cat_filter[0]

    return [[item["id"], item["name"], 99] for item in cat_filter["items"]]


async def parse(sheet, cat):
    """We don't have to transform data into a TreeType, since\n
    categories JSON is already representing it. Also dictionaries\n
    are mutable, so it's the same as managing pointers, we don't
    copy anything."""

    # traverse category tree DFS way
    node_stack = deque([(cat, 1)])
    result = []
    while node_stack:
        node_t = node_stack.pop()
        cat = node_t[0]
        dlevel = node_t[1]

        # write category info
        result.append([cat["id"], cat["name"], dlevel])

        # is list category
        if not cat.get("childs"):
            subjects = await parse_filters(cat)
            result.extend(subjects)
            continue

        # not a list category -> push children and traverse further
        for ch_cat in reversed(cat["childs"]):
            node_stack.append((ch_cat, dlevel + 1))

    # dump data into sheet
    [sheet.append(row) for row in result]


async def main():
    blacklisted = {"Wibes", "Экспресс", "Акции", "Грузовая доставка"}
    cat_list_url = "https://static-basket-01.wbbasket.ru/vol0/data/main-menu-by-ru-v3.json"

    start_time = time.perf_counter()

    wb = openpyxl.Workbook()
    tasks = []
    async with aiohttp.ClientSession() as session:
        async with session.get(cat_list_url) as response:
            for cat in await response.json():
                # skip non-categories
                if cat["name"] in blacklisted:
                    continue
                # initialize category sheet
                sheet = wb.create_sheet(cat["name"])
                sheet.append(["ID", "Name", "Level"])
                tasks.append(parse(sheet, cat))
    # run concurrently for each sheet
    res = await asyncio.gather(*tasks)
    # remove default sheet
    del wb[wb.sheetnames[0]]
    wb.save("test.xlsx")
    logger.info("SUCCESS!")

    end_time = time.perf_counter()
    logger.info(f"Time elapsed: {(end_time - start_time):.2f} seconds.")


if __name__ == "__main__":
    asyncio.run(main())
